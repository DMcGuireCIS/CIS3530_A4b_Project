from flask import Flask, render_template, request, redirect, session, flash, url_for, Response
from werkzeug.security import check_password_hash
from db import get_db_connection
from dotenv import load_dotenv
from io import StringIO
from openpyxl import load_workbook
import os
import psycopg
import csv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ["SECRET_KEY"]

# -----------------------------------------------------------------------
# Flask Session Helpers
# -----------------------------------------------------------------------
def ensure_logged_in():
    """
    Ensures a user is logged in by checking if the username is in the Flask
    session cookie. If not, they are redirected to the login page.
    """
    if "username" not in session:
        return redirect(url_for("login"))
    return None

def current_role():
    """Returns current user role (either admin or viewer)"""
    return session.get("role", None)

def is_admin():
    """Boolean check for if user is admin"""
    return session.get("role") == "admin"

def is_viewer():
    """Boolean check for if user is viewer"""
    return session.get("role") == "viewer"

@app.context_processor
def inject_roles():
    """
    Runs automatically before rendering any template due to @app.context_processor flag
    and allows HTML templates to use the injected roles/functions inside
    """
    return {
        "current_role": current_role,
        "is_admin": is_admin,
        "is_viewer": is_viewer
    }

# -----------------------------------------------------------------------
# A1: Authentication
# -----------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():

    """
    Displays the login page (GET) and processes login forms (POST) through username
    and password authentication.
    """

    # If the user submitted the login form ("POST" request)...
    if request.method == "POST":

        # Set username and password variables to submitted form values
        username = request.form["username"]
        password = request.form["password"]

        # Connect to company database with PostgreSQL...
        with get_db_connection() as connection:
            with connection.cursor() as cursor:

                """
                Gets the username, password_hash, and role of the row corresponding to the
                submitted username. It will either be a row containing (id, password_hash,
                role) if it finds a matching row, or None if no matches are found.

                SQL injection (malicious/dangerous SQL inserted into a form field) is
                prevented by using parameterized SQL. Since a placeholder %s is used,
                the value of username is passed separately in a tuple containing
                parameters (which is (username,) in this case), so PostgreSQL correctly
                treats the username as data and not SQL code that can harm the system.

                """
                cursor.execute("SELECT id, password_hash, role FROM app_user WHERE username = %s", (username,))

                # Fetches the user containing (username, password_hash, role) or None
                user = cursor.fetchone()

        # If no user was found or the password does not match the stored hash...
        if not user or not check_password_hash(user[1], password):

            # Flash error and redirect early back to login
            flash("Invalid username or password.")
            return redirect(url_for("login"))
        
        # Store username in Flask session cookie to mark them as logged in
        session["username"] = user[0]

        # Store role in Flask session cookie to differentiate views
        session["role"] = user[2]

        # Automatically redirect to home page (A2)
        return redirect(url_for("home"))            

    # When the request is GET (not POST) or login fails, show the login form
    return render_template("login.html")

@app.route("/logout")
def logout():

    # Clear the Flask session cookie
    session.clear()

    # Redirect to login page
    return redirect(url_for("login"))

# -----------------------------------------------------------------------
# A2: Home - Employee Overview
# -----------------------------------------------------------------------
@app.route("/")
@app.route("/home")
def home():
    """Employee overview page"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Get filters based on URL
    name_filter = request.args.get("search", "").strip()
    dept_filter = request.args.get("dept", "").strip()
    dept_filter = None if dept_filter == "" else int(dept_filter)
    sort = request.args.get("sort", "name_asc")

    # Whitelisted sorting options (prevents SQL injection)
    sort_options = {
        "name_asc":     "E.Lname ASC, E.Fname ASC",
        "name_desc":    "E.Lname DESC, E.Fname DESC",
        "hours_asc":    "total_hours ASC",
        "hours_desc":   "total_hours DESC"
    }
    order = sort_options.get(sort, "E.Lname ASC")

    """
    Main employee overview query.
    Required columns:
    1. Employee full name
    2. Department name (from Department)
    3. Number of dependents (LEFT JOIN for those with 0 dependents)
    4. Number of projects (LEFT JOIN for those with 0 projects)
    5. Total hours (LEFT JOIN for those with 0 hours)
    COALESCE is used for requirements 3-5 so that the counts and sums still show up as 0.
    """

    sql = f"""
        SELECT 
            E.Ssn,

            -- Build full name from first name, middle initial, and last name
            E.Fname || ' ' || COALESCE(E.Minit || ' ', '') || E.Lname AS full_name,

            D.Dname AS department_name,
            COALESCE(dep.dependent_count, 0) AS dependent_count,
            COALESCE(w.project_count, 0) AS project_count,
            COALESCE(w.total_hours, 0) AS total_hours
        FROM Employee E
        LEFT JOIN Department D 
            ON E.Dno = D.Dnumber

        -- Count dependents per employee and left join to table
        LEFT JOIN (
            SELECT Essn, COUNT(*) AS dependent_count
            FROM Dependent
            GROUP BY Essn
        ) dep 
            ON dep.Essn = E.Ssn

        -- Count projects + sum hours per employee and left join to table
        LEFT JOIN (
            SELECT Essn, COUNT(*) AS project_count, SUM(Hours) AS total_hours
            FROM Works_On
            GROUP BY Essn
        ) w
            ON w.Essn = E.Ssn

        /* 
        Applies department filter and case insensitive match on employee name.

        string = '' is used for when no option is selected/no name is inputted so that the
        condition always evaluates to true and selects everything. For name, it uses
        ILIKE to match partial case-insensitive on a combined string of first and last
        name. For department, it just matches the employee department number.
        */
        WHERE (%s = '' OR (E.Fname || ' ' || E.Lname) ILIKE %s)
          AND (%s::integer IS NULL OR E.Dno = %s::integer)

        -- Orders by selected order type
        ORDER BY {order}
    """

    """
    In matching for ILIKE, since SQL treats % as a wildcard meaning any sequence of
    characters, name pattern is set to start and end with %.
    """
    name_pattern = f"%{name_filter}%"

    # Connect to company database with PostgreSQL...
    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            
            # Execute the SQL with the correct parameters for the filters
            cursor.execute(sql, (name_filter, name_pattern, dept_filter, dept_filter)) # type: ignore[arg-type]

            # Fetch the matched employee list with relevant information as defined in the SQL
            employees = cursor.fetchall()

            # Get department list
            cursor.execute("SELECT Dnumber, Dname FROM Department ORDER BY Dname;")
            departments = cursor.fetchall()

    return render_template(
        "home.html",
        employees=employees,
        departments=departments,
        search=name_filter,
        dept_filter=dept_filter,
        sort=sort
    )

# -----------------------------------------------------------------------
# A3: Projects - Portfolio Summary
# -----------------------------------------------------------------------
@app.route("/")
@app.route("/projects")
def projects():
    """Project overview page"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Whitelisted sorting options (prevents SQL injection)
    sort = request.args.get("sort", "name_asc")
    sort_options = {
        "headcount_asc": "headcount ASC",
        "headcount_desc": "headcount DESC",
        "hours_asc": "total_hours ASC",
        "hours_desc": "total_hours DESC",
    }
    order = sort_options.get(sort, "p.Pname ASC")

    """
    Project portfolio summary query.
    Required columns:
    1. Project Name
    2. Owning Department Name
    3. Headcount
    4. Total Assigned Hours
    COALESCE to maintain projects with 0 employees, headcount and total hours
    """
    sql = f"""
        SELECT 
            p.Pnumber,
            p.Pname,
            d.Dname AS department_name,
            COALESCE(a.headcount, 0) AS headcount,
            COALESCE(a.total_hours, 0) AS total_hours
        FROM Project p
        LEFT JOIN Department d ON p.Dnum = d.Dnumber
        LEFT JOIN AllProjectsWithHeadcount a ON a.Pnumber = p.Pnumber
        ORDER BY {order};
    """

    # Connect to company database with PostgreSQL...
    with get_db_connection() as connection:
        with connection.cursor() as cursor:

            # Get project list
            cursor.execute(sql) # type: ignore
            projects = cursor.fetchall()

            # Get department list
            cursor.execute("SELECT Dnumber, Dname FROM Department ORDER BY Dname;")
            departments = cursor.fetchall()

    return render_template(
        "projects.html",
        projects=projects,
        departments=departments,
        sort=sort
    )

# -----------------------------------------------------------------------
# A4: Project Details - Details & Assignment "Upsert"
# -----------------------------------------------------------------------
@app.route("/project/<int:project_id>")
def project_details(project_id):
    """Specific project details page"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Connect to company database with PostgreSQL...
    with get_db_connection() as connection:
        with connection.cursor() as cursor:

            # Get project
            cursor.execute("""
                SELECT Pnumber, Pname, Dname, Plocation
                FROM Project
                JOIN Department ON Dnumber = Dnum
                WHERE Pnumber = %s;
            """, (project_id,))
            project = cursor.fetchone()

            # Get employees working on project
            cursor.execute("""
                SELECT 
                    e.Ssn,
                    e.Fname || ' ' || COALESCE(e.Minit || ' ', '') || e.Lname AS full_name,
                    d.Dname,
                    w.Hours
                FROM Works_On w
                JOIN Employee e ON e.Ssn = w.Essn
                JOIN Department d ON d.Dnumber = e.Dno
                WHERE w.Pno = %s
                ORDER BY full_name;
            """, (project_id,))
            employees = cursor.fetchall()

            # Get all employees for dropdown
            cursor.execute("""
                SELECT 
                    Ssn,
                    Fname || ' ' || COALESCE(Minit || ' ', '') || Lname AS full_name
                FROM Employee
                ORDER BY full_name;
            """)
            all_employees = cursor.fetchall()

    return render_template(
        "project_details.html",
        project=project,
        employees=employees,
        all_employees=all_employees
    )

@app.route("/project/<int:project_id>/add", methods=["POST"])
def add_assignment(project_id):
    """Updates employee hours"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Redirect to project details page if not admin
    if not is_admin():
        flash("Only admin users may modify assignments.")
        return redirect(url_for("project_details", project_id=project_id))

    # Get submitted form values
    employee_ssn = request.form["employee_ssn"]
    hours = float(request.form["hours"])

    # Connect to company database with PostgreSQL...
    with get_db_connection() as connection:
        with connection.cursor() as cursor:

            # Update hours of employee with matching SSN
            cursor.execute("""
                INSERT INTO Works_On (Essn, Pno, Hours)
                VALUES (%s, %s, %s)
                ON CONFLICT (Essn, Pno)
                DO UPDATE SET Hours = Works_On.Hours + EXCLUDED.Hours;
            """, (employee_ssn, project_id, hours))

    # Redirect to project overview page
    return redirect(url_for("project_details", project_id=project_id))

# -----------------------------------------------------------------------
# A5: Employee Managment (CRUD)
# -----------------------------------------------------------------------
@app.route("/employees")
def employees():
    """Base employee list page"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Connect to company database with PostgreSQL...
    with get_db_connection() as connection:
        with connection.cursor() as cursor:

            # Get (employee SSN, name, address, salary, department name) list
            cursor.execute("""
                SELECT 
                    E.Ssn,
                    E.Fname || ' ' || COALESCE(E.Minit || ' ', '') || E.Lname AS full_name,
                    E.Address,
                    E.Salary,
                    D.Dname
                FROM Employee E
                LEFT JOIN Department D ON E.Dno = D.Dnumber
                ORDER BY full_name;
            """)
            employees = cursor.fetchall()

    return render_template("employees.html", employees=employees)

@app.route("/employee/add", methods=["GET", "POST"])
def add_employee():
    """Add employee page"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Redirect to base employee list page if not admin
    if not is_admin():
        flash("Only admin users may add employees.")
        return redirect(url_for("employees"))

    # Connect to company database with PostgreSQL...
    with get_db_connection() as connection:
        with connection.cursor() as cursor:

            # Get department list
            cursor.execute("SELECT Dnumber, Dname FROM Department ORDER BY Dname;")
            departments = cursor.fetchall()

    # If user adds employee...
    if request.method == "POST":

        # Get submitted form values
        ssn = request.form["ssn"]
        fname = request.form["fname"]
        minit = request.form["minit"] or None
        lname = request.form["lname"]
        address = request.form["address"]
        sex = request.form["sex"]
        salary = request.form["salary"]
        dno = request.form["dno"]

        # Enter try block for primary key collision exception
        try:

            # Connect to company database with PostgreSQL...
            with get_db_connection() as connection:
                with connection.cursor() as cursor:

                    # Insert employee with specified values (avoid SQL injection)
                    cursor.execute("""
                        INSERT INTO Employee (Fname, Minit, Lname, Ssn, Address, Sex, Salary, Dno)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s);
                    """, (fname, minit, lname, ssn, address, sex, salary, dno))

            # Redirect to base employee list page
            return redirect(url_for("employees"))

        # Any psycopg error handled here
        except psycopg.Error:
            flash("Error adding employee. SSN may already exist.")
            return redirect(url_for("add_employee"))

    return render_template("employee_add.html", departments=departments)

@app.route("/employee/<ssn>/edit", methods=["GET", "POST"])
def edit_employee(ssn):
    """Edit employee page"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Redirect to base employee list page if not admin
    if not is_admin():
        flash("Only admin users may edit employees.")
        return redirect(url_for("employees"))

    # Connect to company database with PostgreSQL...
    with get_db_connection() as connection:
        with connection.cursor() as cursor:

            # Get employee with matching SSN
            cursor.execute("""
                SELECT Address, Salary, Dno
                FROM Employee
                WHERE Ssn = %s;
            """, (ssn,))
            employee = cursor.fetchone()

            # Get department number and name list
            cursor.execute("""
                SELECT Dnumber, Dname
                FROM Department
                ORDER BY Dname;
            """)
            departments = cursor.fetchall()

    # If user edits employee...
    if request.method == "POST":

        # Get submitted form values
        address = request.form["address"]
        salary = request.form["salary"]
        dno = request.form["dno"]

        # Connect to company database with PostgreSQL...
        with get_db_connection() as connection:
            with connection.cursor() as cursor:

                # Update employee with matching SSN to form values
                cursor.execute("""
                    UPDATE Employee
                    SET Address = %s,
                        Salary = %s,
                        Dno = %s
                    WHERE Ssn = %s;
                """, (address, salary, dno, ssn))

        # Redirect to base employee list page
        return redirect(url_for("employees"))

    return render_template("employee_edit.html", ssn=ssn, employee=employee, departments=departments)

@app.route("/employee/<ssn>/delete", methods=["POST"])
def delete_employee(ssn):
    """Delete employee"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Redirect to base employee list page if not admin
    if not is_admin():
        flash("Only admin users may delete employees.")
        return redirect(url_for("employees"))

    # Enter try block for foreign key constraint errors
    try:

        # Connect to company database with PostgreSQL...
        with get_db_connection() as connection:
            with connection.cursor() as cursor:

                # Delete employee with matching SSN
                cursor.execute("DELETE FROM Employee WHERE Ssn = %s;", (ssn,))

        # Redirect to base employee list page
        return redirect(url_for("employees"))

    # Any psycopg error handled here
    except psycopg.Error:
        flash("Cannot delete employee: They are assigned to projects, have dependents listed, or are a manager/supervisor.")
        return redirect(url_for("employees"))

# -----------------------------------------------------------------------
# A6: Managers Overview
# -----------------------------------------------------------------------
@app.route("/managers")
def managers():
    """Managers page"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Connect to company database with PostgreSQL...
    with get_db_connection() as connection:
        with connection.cursor() as cursor:

            """
            Get managers list:
            1) Join manager SSN to their employee data (SSNs equal)
            2) Join employees to their department
            3) Join employees to works on data
            3) Group legibly
            """
            cursor.execute("""
                SELECT
                    D.Dname,
                    D.Dnumber,
                    COALESCE(
                        M.Fname || ' ' || COALESCE(M.Minit || ' ', '') || M.Lname,
                        'N/A'
                    ) AS manager_name,
                    COUNT(E.Ssn) AS emp_count,
                    COALESCE(SUM(W.Hours), 0) AS total_hours
                FROM Department D
                LEFT JOIN Employee M ON D.Mgr_ssn = M.Ssn
                LEFT JOIN Employee E ON E.Dno = D.Dnumber
                LEFT JOIN Works_On W ON W.Essn = E.Ssn
                GROUP BY D.Dname, D.Dnumber, manager_name
                ORDER BY D.Dnumber;
            """)
            managers = cursor.fetchall()

    return render_template("managers.html", managers=managers)

# -----------------------------------------------------------------------
# Bonus Feature: CSV Export
# -----------------------------------------------------------------------
@app.route("/home/export")
def home_export():
    """Exports employee overview"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Get filters based on URL
    name_filter = request.args.get("search", "").strip()
    dept_filter = request.args.get("dept", "").strip()
    dept_filter = None if dept_filter == "" else int(dept_filter)
    sort = request.args.get("sort", "name_asc")

    # Whitelisted sorting options (prevents SQL injection)
    sort_options = {
        "name_asc":     "E.Lname ASC, E.Fname ASC",
        "name_desc":    "E.Lname DESC, E.Fname DESC",
        "hours_asc":    "total_hours ASC",
        "hours_desc":   "total_hours DESC"
    }
    order = sort_options.get(sort, "E.Lname ASC")

    # See explanation comment in def home()
    sql = f"""
        SELECT 
            E.Ssn,
            E.Fname || ' ' || COALESCE(E.Minit || ' ', '') || E.Lname AS full_name,
            D.Dname AS department_name,
            COALESCE(dep.dependent_count, 0) AS dependent_count,
            COALESCE(w.project_count, 0) AS project_count,
            COALESCE(w.total_hours, 0) AS total_hours
        FROM Employee E
        LEFT JOIN Department D 
            ON E.Dno = D.Dnumber
        LEFT JOIN (
            SELECT Essn, COUNT(*) AS dependent_count
            FROM Dependent
            GROUP BY Essn
        ) dep 
            ON dep.Essn = E.Ssn
        LEFT JOIN (
            SELECT Essn, COUNT(*) AS project_count, SUM(Hours) AS total_hours
            FROM Works_On
            GROUP BY Essn
        ) w
            ON w.Essn = E.Ssn
        WHERE (%s = '' OR (E.Fname || ' ' || E.Lname) ILIKE %s)
          AND (%s::integer IS NULL OR E.Dno = %s::integer)
        ORDER BY {order};
    """

    # See explanation comment in def home()
    name_pattern = f"%{name_filter}%"

    # Connect to company database with PostgreSQL...
    with get_db_connection() as connection:
        with connection.cursor() as cursor:
            
            # Execute the SQL with the correct parameters for the filters
            cursor.execute(sql, (name_filter, name_pattern, dept_filter, dept_filter)) # type: ignore[arg-type]

            # Get employee list
            employees = cursor.fetchall()

    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)

    # Columns
    writer.writerow(["SSN",
                     "Name",
                     "Department",
                     "Dependents",
                     "Projects",
                     "Total Hours"]
                     )

    # Rows
    for employee in employees:
        writer.writerow(employee)

    # Move cursor back to beginning of file
    output.seek(0)

    # Return downloadable CSV file
    return Response(
        output.read(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=employee_overview.csv"}
    )

@app.route("/projects/export")
def projects_export():
    """Exports project overview"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Whitelisted sorting options (prevents SQL injection)
    sort = request.args.get("sort", "name_asc")
    sort_options = {
        "headcount_asc": "headcount ASC",
        "headcount_desc": "headcount DESC",
        "hours_asc": "total_hours ASC",
        "hours_desc": "total_hours DESC",
    }
    order = sort_options.get(sort, "p.Pname ASC")

    # See explanation comment in def projects()
    sql = f"""
        SELECT 
            p.Pnumber,
            p.Pname,
            d.Dname AS department_name,
            COALESCE(a.headcount, 0) AS headcount,
            COALESCE(a.total_hours, 0) AS total_hours
        FROM Project p
        LEFT JOIN Department d ON p.Dnum = d.Dnumber
        LEFT JOIN AllProjectsWithHeadcount a ON a.Pnumber = p.Pnumber
        ORDER BY {order};
    """

    # Connect to company database with PostgreSQL...
    with get_db_connection() as connection:
        with connection.cursor() as cursor:

            # Get project list
            cursor.execute(sql) # type: ignore
            projects = cursor.fetchall()

    # Create CSV in memory
    output = StringIO()
    writer = csv.writer(output)

    # Columns
    writer.writerow(["Project Number",
                     "Project Name",
                     "Owning Department",
                     "Headcount",
                     "Total Assigned Hours"]
                     )

    # Rows
    for project in projects:
        writer.writerow(project)

    # Move cursor back to beginning of file
    output.seek(0)

    # Return downloadable CSV file
    return Response(
        output.read(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=project_overview.csv"}
    )

# -----------------------------------------------------------------------
# Bonus Feature: Excel Import
# -----------------------------------------------------------------------
def parse_employee_sheet(sheet):
    """
    Parses .xlsx file for Employee data
    Columns: Fname, Minit, Lname, Ssn, Address, Sex, Salary, Super_ssn, Dno, BDate, EmpDate
    """

    # Iterate and take the values of only the first row of the sheet to get column names
    cols = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    expected = ["Fname",
                "Minit",
                "Lname",
                "Ssn",
                "Address",
                "Sex",
                "Salary",
                "Super_ssn",
                "Dno",
                "BDate",
                "EmpDate"]

    # Validate that each column (either col value or "") matches the expected list
    if [col or "" for col in cols] != expected:
        raise ValueError("Header must be exactly: " + ", ".join(expected))
    employees = []

    # Iterate and take the values of every employee from row 2 onwards...
    for idx, employee in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

        # Unpack all values into corresponding variable names
        (fname, minit, lname, ssn, address, sex, salary, super_ssn, dno, bdate, empdate) = employee

        # Validate required fields (see in company_v3.02.sql)
        if not (fname and minit and lname and ssn and address and sex and salary and dno):
            raise ValueError(f"Missing required field on row {idx}.")

        # Validate middle initial is exactly 1 character
        if len(str(minit)) != 1:
            raise ValueError(f"Minit must be 1 character on row {idx}.")

        # Validate sex is exactly 1 character
        if len(str(sex)) != 1:
            raise ValueError(f"Sex must be 1 character on row {idx}.")

        # Validate salary is an integer
        try:
            salary = int(salary)
        except:
            raise ValueError(f"Salary must be integer on row {idx}.")

        # Validate dno is an integer
        try:
            dno = int(dno)
        except:
            raise ValueError(f"Dno must be integer on row {idx}.")

        # Set super_ssn to None if no value is there
        super_ssn = super_ssn or None

        # Append employee to employee list
        employees.append((fname, minit, lname, ssn, address, sex, salary, super_ssn, dno, bdate, empdate))

    return employees

@app.route("/employees/import", methods=["GET", "POST"])
def import_employees():
    """Import employee page"""

    # Redirect to login page if user is not authenticated
    login_redirect = ensure_logged_in()
    if login_redirect: return login_redirect

    # Redirect to base employee list page if not admin
    if not is_admin():
        flash("Only admin users may import employees.")
        return redirect(url_for("employees"))

    # If GET, just render page
    if request.method == "GET":
        return render_template("import_employees.html")

    # If import (POST), perform import...

    # Get file from HTML form
    file = request.files.get("file")

    # Validate file existence and filename
    if not file or not file.filename:
        flash("Please upload a file.")
        return redirect(url_for("import_employees"))
    
    # Validate file type (.xlsx)
    if not file.filename.lower().endswith(".xlsx"):
        flash("Please upload a valid .xlsx file.")
        return redirect(url_for("import_employees"))

    # Load xlsx workbook
    try:
        workbook = load_workbook(file, data_only=True) # type: ignore[arg-type]
        sheet = workbook.active
    except:
        flash("Unable to read Excel file.")
        return redirect(url_for("import_employees"))

    # Parse and validate employees
    try:
        employees = parse_employee_sheet(sheet)
    except ValueError as e:
        flash(str(e))
        return redirect(url_for("import_employees"))

    # Insert employees transactionally (meaning any previous inserts are reversed if one error occurs)
    try:

        # Connect to company database with PostgreSQL...
        with get_db_connection() as connection:
            with connection.cursor() as cursor:

                # Insert every employee into Employee table
                for employee in employees:
                    cursor.execute("""
                        INSERT INTO Employee
                        (Fname, Minit, Lname, Ssn, Address, Sex, Salary,
                         Super_ssn, Dno, BDate, EmpDate)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """, employee)

        flash(f"Successfully imported {len(employees)} employees.")
        return redirect(url_for("employees"))

    except Exception as e:
        flash(f"Import failed (maybe duplicate SSN): {str(e)}")
        return redirect(url_for("import_employees"))

if __name__ == "__main__":
    app.run(debug=True)